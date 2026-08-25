from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from atip_backend.models.destination import Destination


class DestinationImporter:
    """Read, validate, and import destination data from Excel files."""

    REQUIRED_COLUMNS = {
        "Place Name",
        "Built By",
        "Year Built",
        "Location / Distance",
        "Description",
        "Source / Verification",
    }

    def _clean_string(self, value):
        """Normalize optional string values."""

        if value is None:
            return None

        value = str(value).strip()

        return value or None

    def read_excel(self, file_path: str | Path) -> list[dict]:
        """Read and validate destination records from an Excel workbook."""

        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(
                f"Excel file not found: {file_path}"
            )

        workbook = load_workbook(
            file_path,
            read_only=True,
            data_only=True,
        )

        if "Bahawalpur" not in workbook.sheetnames:
            raise ValueError(
                "Required worksheet 'Bahawalpur' was not found."
            )

        worksheet = workbook["Bahawalpur"]

        header_row = None

        for row_index, row in enumerate(
            worksheet.iter_rows(values_only=True),
            start=1,
        ):
            values = [
                self._clean_string(value)
                for value in row
            ]

            if self.REQUIRED_COLUMNS.issubset(set(values)):
                header_row = row_index
                headers = values
                break

        if header_row is None:
            raise ValueError(
                "Could not find the required destination headers."
            )

        records: list[dict] = []

        for row in worksheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ):
            row_data = dict(zip(headers, row))

            name = self._clean_string(
                row_data.get("Place Name")
            )

            if not name:
                continue

            description = self._clean_string(
                row_data.get("Description")
            )

            if not description:
                raise ValueError(
                    f"Destination '{name}' has no description."
                )

            records.append(
                {
                    "name": name,
                    "built_by": self._clean_string(
                        row_data.get("Built By")
                    ),
                    "year_built": self._clean_string(
                        row_data.get("Year Built")
                    ),
                    "location_text": self._clean_string(
                        row_data.get("Location / Distance")
                    ),
                    "description": description,
                    "source_verification": self._clean_string(
                        row_data.get("Source / Verification")
                    ),
                }
            )

        if not records:
            raise ValueError(
                "No valid destination records were found."
            )

        return records

    def check_existing(
        self,
        db: Session,
        records: list[dict],
    ) -> list[str]:
        """Return destination names that already exist in the database."""

        names = [
            record["name"]
            for record in records
        ]

        statement = (
            select(Destination.name)
            .where(Destination.name.in_(names))
        )

        existing_names = db.execute(
            statement
        ).scalars().all()

        return list(existing_names)

    def import_records(
        self,
        db: Session,
        records: list[dict],
    ) -> dict:
        """
        Import validated destination records.

        Existing destination names are skipped so the
        import can safely be run again.
        """

        existing_names = set(
            self.check_existing(
                db,
                records,
            )
        )

        inserted = 0
        skipped = 0

        try:
            for record in records:

                if record["name"] in existing_names:
                    skipped += 1
                    continue

                destination = Destination(
                    name=record["name"],
                    description=record["description"],
                    location_text=record["location_text"],
                    built_by=record["built_by"],
                    year_built=record["year_built"],
                    source_verification=record[
                        "source_verification"
                    ],
                )

                db.add(destination)

                inserted += 1

            db.commit()

        except Exception:
            db.rollback()
            raise

        return {
            "inserted": inserted,
            "skipped": skipped,
            "total": len(records),
        }

    def backfill_geography(
        self,
        db: Session,
        *,
        city: str,
        province: str,
        country: str,
    ) -> int:
        """
        Populate geographic information for destinations
        that do not already have it.
        """

        statement = (
            select(Destination)
            .where(Destination.city.is_(None))
        )

        destinations = db.execute(
            statement
        ).scalars().all()

        updated = 0

        try:
            for destination in destinations:

                destination.city = city
                destination.province = province
                destination.country = country

                updated += 1

            db.commit()

        except Exception:
            db.rollback()
            raise

        return updated